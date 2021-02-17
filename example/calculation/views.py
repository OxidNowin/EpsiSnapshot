from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.utils import timezone

from openpyxl import load_workbook
import numpy as np

from .forms import PostForm
from .models import Post

@login_required
def dashboard(request):
    context = {'selection':'dashboard'}
    return render(request,'calculation/dashboard.html', context)

@login_required
def detail(request, post_id):
    post = get_object_or_404(Post, pk=post_id)
    return render(request, 'calculation/detail.html', {'post': post})

@login_required
def archive(request):
    archive_latest_post_list = Post.objects.order_by('-created_date').filter(author=request.user)
    context = {'archive_latest_post_list': archive_latest_post_list}
    return render(request, 'calculation/archive_index.html', context)

@login_required
def new(request):
    if request.method == "POST":
        form = PostForm(request.POST, request.FILES) 
        try:
            table_of_values = create_table_of_values(request.FILES['file'])
        except:
            form = PostForm()
            return render(request, 'calculation/new.html', {'form': form,
                                                            'error_message':'Упс! Выберете файл с расширением .xlsx,.xlsm,.xltx,.xltm'
                                                            })    
        else:
            if validate(table_of_values) != True:
                form = PostForm()
                return render(request, 'calculation/new.html', {'form': form,
                                                            'error_message':'Упс! Похоже в вашем файле есть некорректные символы или значения ячеек'
                                                            })
            else:
                try:
                    post = form.save(commit=False)
                    post.author = request.user
                    post.created_date = timezone.now()
                    post.file_name = post.file.name
                    table = handle_uploaded_file(table_of_values)
                    post.csi, post.loyalty = count_csi_and_loyalty(table)
                    [post.one,post.two,post.three,post.four,post.five] = count_group(table)
                    post.regressa = regress(table)
                    post.nps = round(nps_calc(table),1) 
                    post.barrier = round(post.loyalty-post.csi,1)
                    post.save()
                except:
                    form = PostForm()
                    return render(request, 'calculation/new.html', {'form': form,
                                                                'error_message':'Упс! Что-то пошло не так'})
                else:
                    return render(request, 'calculation/detail.html', {'post': post})
        
    else:
        form = PostForm()
    return render(request, 'calculation/new.html', {'form': form})

def is_number(str):
    try:
        if float(str)>=1 and float(str)<=10:
            return True
        else:
            return False
    except ValueError:
        return False

def create_table_of_values(file):
    url = file
    table = []
    wb = load_workbook(filename=url, data_only=True, read_only = True)
    ws = wb.active
    table = [row for row in ws.iter_rows(min_row=1, max_col=6, max_row=ws.max_row, values_only=True)]
    return table

def validate(table):
    for column in range(1, len(table)):
        for row in range(0, len(table[column])):
            if (table[column][row] is None) or (is_number(table[column][row])==False):
                return False
    return True


def handle_uploaded_file(table):
    new_table = []
    for column in range(1, len(table)):
        new_table.append([])
        for row in range(0, len(table[column])):
            new_table[column-1].append(table[column][row])
    return new_table
    
def count_csi_and_loyalty(table):
    csi = 0
    loyalty = 0
    new_row = [0.0]*len(table[0])
    for row in range(0, len(new_row)):
        count = 0
        quan = 0
        for column in range(0, len(table)):
            count += float(table[column][row])
            quan += 1
        new_row[row] = round(count/quan, 1)
    csi += (new_row[0]+new_row[1]+new_row[-1])
    csi = 100*(1-((10-csi/3)/9))
    loyalty += (new_row[2]+new_row[3]+new_row[4])
    loyalty = 100*(1-((10-loyalty/3)/9))
    return round(csi,1), round(loyalty,1)

def count_group(table):
    one = 0
    two = 0
    three = 0
    four = 0
    five = 0
    for r in range(0,len(table)):
        csi = table[r][0]+table[r][1]+table[r][-1]
        csi = 100*(1-((10-csi/3)/9))
        loy = table[r][2]+table[r][4]+table[r][5]
        loy = 100*(1-((10-loy/3)/9))
        if csi <= 55:
            if loy <= 55:
                one += 1
            else:
                three += 1
        else:
            if loy <= 55:
                two += 1
        if (loy>=75) and (csi>=75):
            four += 1
    five = len(table)-2-one-two-three-four
    one = round(one/len(table)*100,1)
    two = round(two/len(table)*100,1)
    three = round(three/len(table)*100,1)
    four = round(four/len(table)*100,1)
    five = round(five/len(table)*100,1)
    return [one,two,three,four,five]

def regress(table):
    x_column = []
    for c in range(0,len(table)):
        s = 0
        x_case = []
        s += table[c][0]+table[c][1]+table[c][-1]
        x_case.append(round(s/3,3))
        x_column.append(x_case)
    y_column = []
    for c in range(0,len(table)):
        s = 0
        y_case = []
        s += table[c][2]+table[c][3]+table[c][4]
        y_case.append(round(s/3,3))
        y_column.append(y_case)
    #skm = lm.LinearRegression()
    #skm.fit(x_column, y_column)
    x = np.array(x_column)
    y = np.array(y_column)
    n = np.size(x)
    m_x, m_y = np.mean(x), np.mean(y)
    SS_xy = np.sum(y*x) - n*m_y*m_x
    SS_xx = np.sum(x*x) - n*m_x*m_x
    b_1 = SS_xy / SS_xx    
    return (round(b_1,2))


def nps_calc(f):
    table = f
    crit = 0
    prom = 0
    for row in range(0, len(table[0])):
        for column in range(1, len(table)):
            if table[column][row]<=6:
                crit+=1
            elif table[column][row]>=9:
                prom+=1
    crit /= (len(table[0])*(len(table)-1))
    crit *= 100
    crit = round(crit,2)
    prom /= (len(table[0])*(len(table)-1))
    prom *= 100
    prom = round(prom,2)
    res = prom - crit
    return round(res,2)
